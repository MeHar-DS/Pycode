from openpyxl import Workbook

wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
# ws['A1'] = "Mervin is a Python Expert"
# wb.save("C:\\Mervin\\write_excel.xlsx")  # The workbook is to be saved and not the ws kindly note

for i in range(1,11):
    for j in range(1,6):
        ws.cell(row=i, column=j).value = i+j

wb.save("C:\\Mervin\\write_excel.xlsx")