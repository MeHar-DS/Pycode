from openpyxl import Workbook, load_workbook

w = load_workbook(filename="C:\\Mervin\\write_excel.xlsx")
sh = w.active
sh = w["Sheet"]
# print(sh['A1'].value)

# Another method
# print(w["Sheet"]["A4"].value)

# Another method
# print(sh.cell(row=2, column=4).value)

row_cnt = sh.max_row
col_cnt = sh.max_column

# print("max rows are: {0} and max columns are : {1}".format(row_cnt,col_cnt))

for i in range(1,row_cnt+1):
    for j in range(1,col_cnt+1):
        print(sh.cell(row=i,column=j).value, end=' ')
    print("\n")
