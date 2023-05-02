from faker import Faker
from openpyxl import Workbook

fake_data = Faker(['it_IT', 'en_US',"hi_IN","ja_JP"])  # A list of locales have been specified

# print(fake_data.name())
# print(fake_data.email())
# print(fake_data.address())

w = Workbook()

s = w.active

for i in range(1, 31):
    for j in range(1, 4):
        s.cell(row=i,column= 1).value = fake_data.name()
        s.cell(row=i, column=2).value = fake_data.email()
        s.cell(row=i, column=3).value = fake_data.address()

w.save("testdata.xlsx")